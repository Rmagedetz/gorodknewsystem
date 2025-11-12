import datetime

import pandas as pd
import streamlit as st
import sql
from sqlalchemy.orm import Session

from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
import copy

# Получение разрешений пользователя
permission = st.session_state.user_permissions['editing_groups'][0]

# Селекторы сезонов, филиалов и групп
left, center, right = st.columns(3)

# Сезоны
seasons = sql.Seasons.get_df()['season_name']
with left:
    season_selector = st.selectbox('Сезон', seasons, key='season_selector')

# Филиалы
filials = sql.Filials.get_df(season_name=season_selector)['filial_name']
with center:
    filial_selector = st.selectbox('Филиал', filials, key='filial_selector')

# Группы
groups = sql.Groups.get_df(season_name=season_selector, filial_name=filial_selector)['group_name']
with right:
    group_selector = st.selectbox('Группа', groups, key='group_selector')

group_data = sql.Groups.get_days_for_group_in_season(season_name=season_selector, group_name=group_selector)

# --- Загрузка данных ---

parents_df = sql.Parent.get_df()
childrens_df = sql.Child.get_df()

records_df = sql.Records.get_df(
    season_name=season_selector,
    filial_name=filial_selector,
    group_name=group_selector
)

visits_df = sql.Visits.get_df(
    season_name=season_selector,
    filial_name=filial_selector,
    group_name=group_selector
)

payments_df = sql.Payments.get_df(
    season_name=season_selector,
    group_name=group_selector
)


ankets = sql.Ankets.get_df()
childs_in_group = records_df.merge(ankets, left_on='child_name', right_on='name', how='left')
childs_in_group = childs_in_group.sort_values(by='child_name', ascending=True)

column_config = {
    'child_name': st.column_config.Column('ФИО',
                                          width='medium',
                                          help='ФИО ребенка'),
    'child_birthday': st.column_config.DateColumn('ДР',
                                                  format='DD.MM.YY',
                                                  help='День рождения',
                                                  width='small'),
    'parent_main_name': st.column_config.Column('ФИО Р',
                                                width='medium',
                                                help='ФИО Родителя'),
    'parent_main_phone': st.column_config.Column('Тел.',
                                                 width=130,
                                                 help='Телефон родителя'),
    'parent_add': st.column_config.Column('ФИО ДР',
                                          width='medium',
                                          help='ФИО другого родственника'),
    'phone_add': st.column_config.Column('Доб. тел.',
                                         width=130,
                                         help='Телефон другого родственника'),
    'leave': st.column_config.Column('У',
                                     width=40,
                                     help='Уходит сам'),
    'addr': st.column_config.Column('Адр.',
                                    width='medium',
                                    help='Адрес'),
    'oms': st.column_config.Column('ОМС',
                                   width=125,
                                   help='Полис'),
    'disease': st.column_config.Column('Заболевания',
                                       width=125,
                                       help='Заболевания'),
    'allergy': st.column_config.Column('Аллергия',
                                       width=125,
                                       help='Аллергия'),
    'other': st.column_config.Column('Травмы',
                                     width=125,
                                     help='Травмы'),
    'physic': st.column_config.Column('Ограничения',
                                      width=125,
                                      help='Ограничения'),
    'swimm': st.column_config.Column('Бассенйн',
                                     width=125,
                                     help='Бассенйн'),
    'jacket_swimm': st.column_config.Column('Нарукавники',
                                            width=125,
                                            help='Нарукавники'),
    'hobby': st.column_config.Column('Хобби',
                                     width='medium',
                                     help='Хобби'),
    'additional_info': st.column_config.Column('Доп данные',
                                               width='medium',
                                               help='Доп данные'),
    'additional_contact': st.column_config.Column('Доп контакт',
                                                  width='medium',
                                                  help='Кто кроме родителя забирает ребенка')
}

# --- Объединение данных ---
if not records_df.empty:
    merged = records_df.merge(parents_df, on='parent_name', how='left', suffixes=('', '_p'))
    merged = merged.merge(childrens_df, on='child_name', how='left', suffixes=('', '_c'))

    # Восстанавливаем корректное имя родителя
    for col in ['parent_name', 'parent_name_x', 'parent_name_y', 'parent_name_p']:
        if col in merged.columns:
            merged['parent_name'] = merged[col]
            break
    else:
        merged['parent_name'] = records_df.get('parent_name', '')

    # Выбираем доступные колонки
    available_cols = [c for c in ['child_name', 'child_age', 'parent_name', 'parent_phone', 'record_status', 'comment']
                      if c in merged.columns]
    merged = merged[available_cols].copy()
    merged = merged.sort_values(by='child_name', ascending=True)
    merged.fillna('', inplace=True)
else:
    # Если нет записей, создаем пустой DataFrame с нужными колонками
    merged = pd.DataFrame(
        columns=['child_name', 'child_age', 'parent_name', 'parent_phone', 'record_status', 'comment'])

# --- Добавляем колонки посещаемости ---
for day in range(1, group_data + 1):
    merged[str(day)] = ' '  # по умолчанию пусто

# --- Наполняем существующими посещениями ---
if not visits_df.empty:
    for _, visit in visits_df.iterrows():
        mask = (merged['child_name'] == visit['child_name']) & (merged['parent_name'] == visit['parent_name'])
        if mask.any():
            day_col = str(visit['day_number'])
            merged.loc[mask, day_col] = visit.get('visit_status', '')

# --- Конфигурация столбцов ---
columns_config = {
    'child_name': 'ФИО Ребенка',
    'child_age': 'Возраст',
    'parent_name': 'ФИО Родителя',
    'parent_phone': 'Тел. родителя',
    'record_status': 'Оплата',
    'comment': 'Комментарий'
}

visit_options = [1, "1Д", "X", "Н", "П", "Б", "В"]
for day in range(1, group_data + 1):
    columns_config[str(day)] = st.column_config.SelectboxColumn(str(day), options=visit_options)

(group_list_tab, visits_tab, childrens_tab, drive_tab,
 payments_tab, locker_list_tab, pool_list_tab, adress_tab) = st.tabs(
    ['📄Список',
     '✅Посещаемость',
     '📄Лист ознакомления',
     '🚌Поездка',
     '💳Оплаты',
     '🔢Список на шкафчики',
     '🏊Бассейн',
     '📍Адреса'])

with group_list_tab:
    show = childs_in_group[['child_name',
                            'child_birthday',
                            'parent_main_name',
                            'parent_main_phone',
                            'parent_add',
                            'phone_add',
                            'leave',
                            'addr',
                            'oms',
                            'disease',
                            'allergy',
                            'other',
                            'physic',
                            'swimm',
                            'jacket_swimm',
                            'hobby']]
    group_df = st.dataframe(show,
                            column_config=column_config,
                            hide_index=True)

with visits_tab:
    # --- Отображение редактора ---
    show = st.data_editor(
        merged,
        column_config=columns_config,
        hide_index=True
    )


    # --- Функция проверки валидности статуса посещения ---
    def is_valid_status(status: str) -> bool:
        status = str(status).strip()
        return bool(status) and not any(bad in status for bad in [' ', 'ов', 'на'])


    # --- Сохранение посещаемости ---
    if st.button("💾 Сохранить посещаемость"):
        session: Session = sql.SessionLocal()
        try:
            inserts = []
            for _, row in show.iterrows():
                parent_name = row.get('parent_name', '').strip()
                if not parent_name:
                    st.warning(f"⛔ Пропущен parent_name у {row.get('child_name')}")
                    continue

                for day in range(1, group_data + 1):
                    visit_status = str(row.get(str(day), '')).strip()
                    if not is_valid_status(visit_status):
                        continue
                    visit_status = visit_status[:20]

                    existing = (
                        session.query(sql.Visits)
                        .filter_by(
                            season_name=season_selector,
                            filial_name=filial_selector,
                            group_name=group_selector,
                            child_name=row['child_name'],
                            parent_name=parent_name,
                            day_number=day
                        )
                        .first()
                    )

                    if existing:
                        existing.visit_status = visit_status
                    else:
                        inserts.append(sql.Visits(
                            season_name=season_selector,
                            filial_name=filial_selector,
                            group_name=group_selector,
                            child_name=row['child_name'],
                            parent_name=parent_name,
                            day_number=day,
                            visit_status=visit_status
                        ))

            if inserts:
                session.bulk_save_objects(inserts)

            session.commit()
            st.success("✅ Посещаемость сохранена!")
            st.rerun()
        except Exception as e:
            session.rollback()
            st.error(f"Ошибка при сохранении: {e}")
        finally:
            session.close()

day_columns = [str(i) for i in range(1, group_data + 1)]


def count_visited(row):
    return sum(row[day] in ['1', '1Д', 'Н'] for day in day_columns)


with payments_tab:
    first_df = merged.copy()
    del first_df['child_age']
    del first_df['parent_phone']
    del first_df['record_status']
    del first_df['comment']

    first_df['days'] = first_df.apply(count_visited, axis=1)
    columns_config['days'] = 'Отхожено дней'

    pivot = payments_df.pivot_table(
        index=["child_name"],
        values="pay_sum",
        aggfunc="sum"
    ).reset_index()
    columns_config['pay_sum'] = "Оплачено"
    merged = first_df.merge(pivot, on='child_name', how='left')

    pivot_options = payments_df.pivot_table(
        index=["child_name"],
        values="option",
        aggfunc="first"
    ).reset_index()
    merged = merged.merge(pivot_options, on='child_name', how='left')
    columns_config['option'] = 'Тариф'

    pivot_2 = payments_df.pivot_table(
        index=["child_name"],
        values="pay_form",
        aggfunc="first"
    ).reset_index()

    merged = merged.merge(pivot_2, on="child_name", how='left')
    pay_forms = list(sql.Payments_forms.get_df()['form'])
    columns_config['pay_form'] = st.column_config.SelectboxColumn('Тип оплаты', options=pay_forms)

    try:
        merged['sum_to_debit'] = merged['pay_sum'] / group_data * merged['days']
    except:
        pass

    pay_forms_options = list(sql.Payments_forms.get_df()['form'])
    columns_config['sum_to_debit'] = 'Сумма к списанию'

    merged['go'] = False
    columns_config['go'] = 'Отметка'

    show2 = st.data_editor(
        merged,
        column_config=columns_config,
        hide_index=True,
        key="sd"
    )

    if st.button('Произвести списания'):
        first_step = show2[show2['go'] == True][['child_name', 'parent_name', 'pay_sum', 'pay_form', 'sum_to_debit']]
        for _, row in first_step.iterrows():
            sql.Debits.add_object(
                datetime=datetime.datetime.now(),
                account=st.session_state.user,
                season_name=season_selector,
                filial_name=filial_selector,
                group_name=group_selector,
                child_name=row['child_name'],
                parent_name=row['parent_name'],
                pay_sum=row['sum_to_debit'],
                pay_form=row['pay_form'],
                option='Путевка',
                comment=''
            )
        st.rerun()

with childrens_tab:
    show = childs_in_group[['child_name',
                            'child_birthday',
                            'disease',
                            'allergy',
                            'physic',
                            'other',
                            'leave',
                            'jacket_swimm',
                            'additional_info',
                            'additional_contact']]
    df = st.dataframe(show, column_config=column_config, hide_index=True)


    def create_ds(df):
        output = BytesIO()

        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("Лист ознакомления")

        # ws.column_dimensions["A"].width = 42

        num_rows = ws.max_row + 4

        row_num = 0
        for _, row in df.iterrows():
            cell = ws.cell(row=num_rows + row_num, column=1, value=row_num + 1)
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=2, value=row['child_name'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            try:
                cell = ws.cell(row=num_rows + row_num, column=3, value=row['child_birthday'].strftime('%d.%m.%Y'))
                cell.font = Font(name='TimesNewRoman', size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            except:
                cell = ws.cell(row=num_rows + row_num, column=3, value="")
                cell.font = Font(name='TimesNewRoman', size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=4, value=row['disease'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=5, value=row['allergy'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=6, value=row['other'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=7, value=row['physic'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=8, value=row['leave'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=9, value=row['jacket_swimm'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=10, value=row['additional_info'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=11, value=row['additional_contact'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            row_num += 1

        num_rows = ws.max_row + 2

        for counter in range(5):
            cell = ws.cell(row=num_rows + counter, column=1, value="Ознакомлен/а __________________  "
                                                                   "______________________/_____________/ “____” ______ "
                                                                   "202__ г.")
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            ws.merge_cells(start_row=num_rows + counter, start_column=1, end_row=num_rows + counter, end_column=11)

        cell = ws.cell(row=1, column=1, value="Лист ознакомления инструкторов группы по уходу и присмотру об "
                                              "особенностях здоровья детей")
        cell.font = Font(name='TimesNewRoman', size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

        dates = sql.Groups.get_start_end(season_selector, filial_selector, group_selector)
        cell = ws.cell(row=2, column=1,
                       value=f"Довожу до вашего сведения особенности здоровья детей группы {group_selector} "
                             f"находящихся в клубе в период с {dates['start_date'].strftime('%d.%m.%Y')} "
                             f"г. по {dates['end_date'].strftime('%d.%m.%Y')} г.")
        cell.font = Font(name='TimesNewRoman', size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
        ws.row_dimensions[2].height = 30

        cell = ws.cell(row=3, column=1,
                       value=f"Евдокимова Е.Б. ___________________")
        cell.font = Font(name='TimesNewRoman', size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)

        cell = ws.cell(row=4, column=1, value=f"№")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=2, value=f"ФИО ребенка")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=3, value=f"Дата рождения")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=4, value=f"Заболевания")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=5, value=f"Аллергия")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=6, value=f"Травмы")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=7, value=f"Ограничения")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=8, value=f"Уходит сам")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=9, value=f"Нарукавники")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=10, value=f"Доп. данные")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=11, value=f"Кто кроме родителя забирает ребенка?")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 17
        ws.column_dimensions["C"].width = 9
        ws.column_dimensions["D"].width = 10

        ws.column_dimensions["E"].width = 9
        ws.column_dimensions["F"].width = 9
        ws.column_dimensions["G"].width = 9
        ws.column_dimensions["H"].width = 9
        ws.column_dimensions["I"].width = 9

        ws.column_dimensions["J"].width = 15
        ws.column_dimensions["K"].width = 15

        page_setup = ws.page_setup
        page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        wb.save(output)
        output.seek(0)
        return output


    if st.button("Скачать список", key="download_list"):
        excel_file = create_ds(show)
        st.download_button(
            label="💾Нажми для скачивания списка",
            data=excel_file,
            file_name=f"Лист ознакомления_{group_selector}_{filial_selector}_{season_selector}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with drive_tab:
    col1, col2 = st.columns(2)
    with col1:
        drive_day = st.date_input('Дата поездки', format='DD.MM.YYYY')
    with col2:
        drive_adr = st.text_input('Адрес поездки')
    show = childs_in_group[['child_name',
                            'child_birthday',
                            'parent_name',
                            'parent_main_phone']]
    df = st.dataframe(show, column_config=column_config, hide_index=True)


    def create_drive_list(df):
        output = BytesIO()

        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("Поездка")

        num_rows = ws.max_row + 4

        row_num = 0
        for _, row in df.iterrows():
            cell = ws.cell(row=num_rows + row_num, column=1, value=row_num + 1)
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=2, value=row['child_name'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            try:
                cell = ws.cell(row=num_rows + row_num, column=3, value=row['child_birthday'].strftime('%d.%m.%Y'))
                cell.font = Font(name='TimesNewRoman', size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            except:
                cell = ws.cell(row=num_rows + row_num, column=3, value="")
                cell.font = Font(name='TimesNewRoman', size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=4, value=row['parent_name'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=5, value=row['parent_main_phone'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            row_num += 1

        num_rows = ws.max_row + 1

        cell = ws.cell(row=num_rows, column=1, value="Сопровождающие")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.merge_cells(start_row=num_rows, start_column=1, end_row=num_rows, end_column=5)

        num_rows = ws.max_row + 1
        cell = ws.cell(row=num_rows, column=1, value="ФИО сопровождающего")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.border = border
        ws.merge_cells(start_row=num_rows, start_column=1, end_row=num_rows, end_column=2)

        cell = ws.cell(row=num_rows, column=3, value="Должность")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.border = border

        cell = ws.cell(row=num_rows, column=4, value="Телефон")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.border = border
        ws.merge_cells(start_row=num_rows, start_column=4, end_row=num_rows, end_column=5)

        num_rows = ws.max_row + 1

        for counter in range(3):
            cell = ws.cell(row=num_rows + counter, column=1, value="_")
            cell.font = Font(name='TimesNewRoman', size=9, bold=True)
            cell.border = border
            ws.merge_cells(start_row=num_rows + counter, start_column=1, end_row=num_rows + counter, end_column=2)

            cell = ws.cell(row=num_rows + counter, column=3, value="_")
            cell.font = Font(name='TimesNewRoman', size=9, bold=True)
            cell.border = border

            cell = ws.cell(row=num_rows + counter, column=4, value="_")
            cell.font = Font(name='TimesNewRoman', size=9, bold=True)
            cell.border = border
            ws.merge_cells(start_row=num_rows + counter, start_column=4, end_row=num_rows + counter, end_column=5)

            ws.row_dimensions[num_rows + counter].height = 30

        num_rows = ws.max_row + 2

        cell = ws.cell(row=num_rows, column=1, value="ИП Евдокимова Е.Б. __________________________")
        cell.font = Font(name='TimesNewRoman', size=9)

        num_rows = ws.max_row + 2

        cell = ws.cell(row=num_rows, column=1, value="Подпись ответственного сопровождающего   "
                                                     "______________/___________________________")
        cell.font = Font(name='TimesNewRoman', size=9)

        cell = ws.cell(row=1, column=1, value="Список пассажиров, которым разрешается (помимо водителей) находиться в "
                                              "автобусе в процессе перевозки")
        ws.row_dimensions[1].height = 30
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

        cell = ws.cell(row=2, column=1, value=f"Дата поездки: {drive_day.strftime('%d.%m.%Y')}")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

        cell = ws.cell(row=2, column=3, value=f"Адрес: {drive_adr}")
        ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)

        cell = ws.cell(row=4, column=1, value=f"№")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=2, value=f"ФИО пассажира")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=3, value=f"Дата рождения")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=4, value=f"ФИО родителя")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=4, column=5, value=f"Телефон")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 20

        wb.save(output)
        output.seek(0)
        return output


    if st.button("Скачать список", key="download_drive"):
        excel_file = create_drive_list(show)
        st.download_button(
            label="💾Нажми для скачивания списка",
            data=excel_file,
            file_name=f"Поездка_{drive_day}_{group_selector}_{filial_selector}_{season_selector}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with locker_list_tab:
    children = list(childs_in_group['child_name'])
    show = st.data_editor(children, hide_index=True, column_config={'value': 'ФИО ребенка'})


    def create_locker_list(chld_lst):
        output = BytesIO()

        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("Шкафчики")

        num_rows = 1

        row_num = 0
        for name in chld_lst:
            cell = ws.cell(row=num_rows + row_num, column=1, value=name)
            cell.font = Font(name='TimesNewRoman', size=30)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            ws.row_dimensions[num_rows + row_num].height = 70

            row_num += 1

        ws.column_dimensions["A"].width = 60

        wb.save(output)
        output.seek(0)
        return output


    if st.button("Скачать список", key="download_lockers"):
        excel_file = create_locker_list(children)
        st.download_button(
            label="💾Нажми для скачивания списка",
            data=excel_file,
            file_name=f"Шкафчики_{group_selector}_{filial_selector}_{season_selector}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with pool_list_tab:
    show = childs_in_group[['child_name',
                            'parent_main_phone',
                            'physic',
                            'swimm',
                            'jacket_swimm']]
    df = st.dataframe(show, column_config=column_config, hide_index=True)


    def create_pool_list(df):
        output = BytesIO()

        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("Бассейн")

        num_rows = ws.max_row + 2

        row_num = 0
        for _, row in df.iterrows():
            cell = ws.cell(row=num_rows + row_num, column=1, value=row_num + 1)
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=2, value=row['child_name'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=3, value=row['parent_main_phone'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=4, value=row['physic'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=5, value=row['swimm'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=6, value=row['jacket_swimm'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            row_num += 1

        num_rows = ws.max_row + 1

        for counter in range(3):
            cell = ws.cell(row=num_rows + counter, column=2,
                           value="Ознакомлен/а __________________  ______________________/_____________/ “____” "
                                 "______ 202__ г.")
            cell.font = Font(name='TimesNewRoman', size=9, bold=True)
            ws.merge_cells(start_row=num_rows + counter, start_column=2, end_row=num_rows + counter, end_column=6)

            ws.row_dimensions[num_rows + counter].height = 30

        num_rows = ws.max_row + 2

        cell = ws.cell(row=2, column=1, value=f"№")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=2, value=f"ФИО Ребенка")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=3, value=f"Телефон")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=4, value=f"Ограничения")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=5, value=f"Бассейн")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=6, value=f"Нарукавники")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=1, column=1, value=f"Лист для посещения бассейна группа {group_selector}")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

        wb.save(output)
        output.seek(0)
        return output


    if st.button("Скачать список", key="download_pool"):
        excel_file = create_pool_list(show)
        st.download_button(
            label="💾Нажми для скачивания списка",
            data=excel_file,
            file_name=f"Бассейн_{group_selector}_{filial_selector}_{season_selector}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with adress_tab:
    show = childs_in_group[['child_name',
                            'addr',
                            'parent_main_phone',
                            'parent_main_name']]
    df = st.dataframe(show, column_config=column_config, hide_index=True)


    def create_addr_list(df):
        output = BytesIO()

        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("Адреса")

        num_rows = ws.max_row + 2

        row_num = 0
        for _, row in df.iterrows():
            cell = ws.cell(row=num_rows + row_num, column=1, value=row_num + 1)
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=2, value=row['child_name'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=3, value=row['addr'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=4, value=row['parent_main_phone'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

            cell = ws.cell(row=num_rows + row_num, column=5, value=row['parent_main_name'])
            cell.font = Font(name='TimesNewRoman', size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            row_num += 1

        num_rows = ws.max_row + 1

        cell = ws.cell(row=2, column=1, value=f"№")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=2, value=f"ФИО Ребенка")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=3, value=f"Адрес проживания")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=4, value=f"Телефон")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=2, column=5, value=f"ФИО Родителя")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        cell = ws.cell(row=1, column=1, value=f"Адреса группа {group_selector}")
        cell.font = Font(name='TimesNewRoman', size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

        wb.save(output)
        output.seek(0)
        return output


    if st.button("Скачать список", key="download_addr"):
        excel_file = create_addr_list(show)
        st.download_button(
            label="💾Нажми для скачивания списка",
            data=excel_file,
            file_name=f"Адреса_{group_selector}_{filial_selector}_{season_selector}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
