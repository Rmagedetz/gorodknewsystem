import pandas as pd
import streamlit as st
import sql

from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
import copy

permission = st.session_state.user_permissions['editing_groups'][0]

ankets_df = sql.Ankets.get_df()

column_config = {'name': '–§–ò–û –†–µ–±–µ–Ω–∫–∞',
                 'parent_main_name': '–§–ò–û –†–æ–¥–∏—Ç–µ–ª—è',
                 'parent_main_phone': '–¢–µ–ª',
                 'child_age': '–í–æ–∑—Ä–∞—Å—Ç',
                 'child_birthday': '–î–†',
                 'parent_add': '–î–æ–ø –∫–æ–Ω—Ç–∞–∫—Ç',
                 'phone_add': '–î–æ–ø –Ω–æ–º–µ—Ä',
                 'leave': '–£—Ö–æ–¥–∏—Ç —Å–∞–º',
                 'additional_contact': '–î–æ–ø –∫–æ–Ω—Ç–∞–∫—Ç',
                 'addr': '–ê–¥—Ä–µ—Å',
                 'disease': '–ó–∞–±–æ–ª–µ–≤–∞–Ω–∏—è',
                 'allergy': '–ê–ª–ª–µ—Ä–≥–∏–∏',
                 'other': '–î—Ä—É–≥–æ–µ',
                 'physic': '–û–≥—Ä–∞–Ω–∏—á–µ–Ω–∏—è',
                 'swimm': '–ë–∞—Å—Å–µ–π–Ω',
                 'jacket_swimm': '–ñ–∏–ª–µ—Ç',
                 'hobby': '–•–æ–±–±–∏',
                 'school': '–®–∫–æ–ª–∞',
                 'additional_info': '–î–æ–ø –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è',
                 'departures': '',
                 'referer': '–û—Ç–∫—É–¥–∞ —É–∑–Ω–∞–ª–∏',
                 'ok': '',
                 'mailing': '–†–∞—Å—Å—ã–ª–∫–∞',
                 'personal_accept': '–°–æ–≥–∞–ª—Å–∏–µ',
                 'oms': '–û–ú–°',
                 'child_balance': '–ë–∞–ª–∞–Ω—Å —Ä–µ–±–µ–Ω–∫–∞',
                 'parent_balance': '–ë–∞–ª–∞–Ω—Å —Ä–æ–¥–∏—Ç–µ–ª—è'}

child_selector = st.selectbox('–†–µ–±–µ–Ω–æ–∫', options=ankets_df['name'])

data = ankets_df[ankets_df['name'] == child_selector].reset_index()

col1, col2, col3 = st.columns(3)
with col1:
    with st.container(border=True, height=600):
        st.subheader('üßí –î–∞–Ω–Ω—ã–µ —Ä–µ–±–µ–Ω–∫–∞')
        st.divider()
        st.write(f'–§–ò–û —Ä–µ–±–µ–Ω–∫–∞: {data['name'][0]}')
        st.write(f'–§–ò–û —Ä–æ–¥–∏—Ç–µ–ª—è: {data['parent_main_name'][0]}')
        st.write(f'–¢–µ–ª–µ—Ñ–æ–Ω –¥–ª—è —Å–≤—è–∑–∏: {data['parent_main_phone'][0]}')
        st.divider()
        st.write(f'–§–ò–û –¥—Ä—É–≥–æ–≥–æ –≤–∑—Ä–æ—Å–ª–æ–≥–æ: {data['parent_add'][0]}')
        st.write(f'–¢–µ–ª–µ—Ñ–æ–Ω –¥–ª—è –∑–∞–ø–∞—Å–Ω–æ–π —Å–≤—è–∑–∏: {data['phone_add'][0]}')

        if data['leave'][0] == '–Ω–µ—Ç':
            st.markdown(f'–£—Ö–æ–¥–∏—Ç —Å–∞–º: :red-background[{data['leave'][0]}]')
        elif data['leave'][0] == '–¥–∞':
            st.markdown(f'–£—Ö–æ–¥–∏—Ç —Å–∞–º: :green-background[{data['leave'][0]}]')
        else:
            st.markdown(f'–£—Ö–æ–¥–∏—Ç —Å–∞–º: :blue-background[{data['leave'][0]}]')

        st.write(f'–ö—Ç–æ –º–æ–∂–µ—Ç –∑–∞–±–∏—Ä–∞—Ç—å: {data['additional_contact'][0]}')
        st.write(f'–ê–¥—Ä–µ—Å –ø—Ä–æ–∂–∏–≤–∞–Ω–∏—è: {data['addr'][0]}')
        st.write(f'–ù–æ–º–µ—Ä –ø–æ–ª–∏—Å–∞: {data['oms'][0]}')

with col2:
    with st.container(border=True, height=600):
        st.subheader('‚öïÔ∏è –î–∞–Ω–Ω—ã–µ –æ –∑–¥–æ—Ä–æ–≤—å–µ')
        st.divider()

        if data['disease'][0] == '–∑–∞–±–æ–ª–µ–≤–∞–Ω–∏–π –Ω–µ—Ç':
            st.markdown(f'–ó–∞–±–æ–ª–µ–≤–∞–Ω–∏—è: :green-background[{data['leave'][0]}]')
        else:
            st.markdown(f'–ó–∞–±–æ–ª–µ–≤–∞–Ω–∏—è: :red-background[{data['disease'][0]}]')

        if data['allergy'][0] == '–Ω–µ—Ç':
            st.markdown(f'–ê–ª–ª–µ—Ä–≥–∏—è: :green-background[{data['allergy'][0]}]')
        else:
            st.markdown(f'–ê–ª–ª–µ—Ä–≥–∏—è: :red-background[{data['allergy'][0]}]')

        if data['other'][0] == '–Ω–µ—Ç':
            st.markdown(f'–¢—Ä–∞–≤–º—ã: :green-background[{data['other'][0]}]')
        else:
            st.markdown(f'–¢—Ä–∞–≤–º—ã: :red-background[{data['other'][0]}]')

        if data['physic'][0] == '–Ω–µ—Ç':
            st.markdown(f'–û—Ä–≥—Ä–∞–Ω–∏—á–µ–Ω–∏—è: :green-background[{data['physic'][0]}]')
        else:
            st.markdown(f'–û—Ä–≥—Ä–∞–Ω–∏—á–µ–Ω–∏—è: :red-background[{data['physic'][0]}]')

with col3:
    with st.container(border=True, height=600):
        st.subheader('üìù –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–æ')
        st.divider()

        if data['swimm'][0] == '–¥–∞':
            st.markdown(f'–ë—É–¥–µ—Ç –ª–∏ –ø–æ—Å–µ—â–∞—Ç—å –±–∞—Å—Å–µ–π–Ω: :green-background[{data['swimm'][0]}]')
        else:
            st.markdown(f'–ë—É–¥–µ—Ç –ª–∏ –ø–æ—Å–µ—â–∞—Ç—å –±–∞—Å—Å–µ–π–Ω: :red-background[{data['swimm'][0]}]')

        if data['jacket_swimm'][0] == '–Ω–µ—Ç':
            st.markdown(f'–ù—É–∂–Ω—ã –ª–∏ –Ω–∞—Ä—É–∫–∞–≤–Ω–∏–∫–∏: :green-background[{data['jacket_swimm'][0]}]')
        else:
            st.markdown(f'–ù—É–∂–Ω—ã –ª–∏ –Ω–∞—Ä—É–∫–∞–≤–Ω–∏–∫–∏: :red-background[{data['jacket_swimm'][0]}]')

        st.write(f'–ß–µ–º —É–≤–ª–µ–∫–∞–µ—Ç—Å—è: {data['hobby'][0]}')
        st.write(f'–®–∫–æ–ª–∞: {data['school'][0]}')
        st.divider()
        st.write(f'–î–æ–ø. –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è: {data['additional_info'][0]}')

with st.container(border=True):
    st.subheader('‚úâÔ∏è –û–±—â–∏–µ –¥–∞–Ω–Ω—ã–µ')
    st.write(f'–ö–∞–∫ —É–∑–Ω–∞–ª–∏ –ø—Ä–æ –ì–æ—Ä–æ–¥–æ–∫: {data['referer'][0]}')
    st.write(f'e-mail: {data['email'][0]}')
    st.write(f'–ü—Ä–æ–≥—É–ª–∫–∏: {data['departures'][0]}')
    st.write(f'–†–∞—Å—Å—ã–ª–∫–∏: {data['mailing'][0]}')
    st.write(f'–°–æ–≥–ª–∞—Å–∏–µ –Ω–∞ –æ–±—Ä–±–æ—Ç–∫—É –ø–µ—Ä—Å–æ–Ω–∞–ª—å–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö: {data['personal_accept'][0]}')


def create_ds(df):
    output = BytesIO()

    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("–ê–Ω–∫–µ—Ç–∞")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–§–ò–û —Ä–µ–±–µ–Ω–∫–∞")
    cell.font = Font(name='TimesNewRoman', size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['name'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–î–∞—Ç–∞ —Ä–æ–∂–¥–µ–Ω–∏—è —Ä–µ–±–µ–Ω–∫–∞")
    cell.font = Font(name='TimesNewRoman', size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['child_birthday'][0].strftime('%d.%m.%Y'))
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–§–ò–û —Ä–æ–¥–∏—Ç–µ–ª—è (–∑–∞–∫–æ–Ω–Ω–æ–≥–æ –ø—Ä–µ–¥—Å—Ç–∞–≤–∏—Ç–µ–ª—è) –¥–ª—è –æ—Å–Ω–æ–≤–Ω–æ–π —Å–≤—è–∑–∏")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['parent_main_name'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–¢–µ–ª–µ—Ñ–æ–Ω –¥–ª—è –æ—Å–Ω–æ–≤–Ω–æ–π —Å–≤—è–∑–∏")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['parent_main_phone'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–§–ò–û –¥—Ä—É–≥–æ–≥–æ –≤–∑—Ä–æ—Å–ª–æ–≥–æ")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['parent_add'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–¢–µ–ª–µ—Ñ–æ–Ω –¥–ª—è –∑–∞–ø–∞—Å–Ω–æ–π —Å–≤—è–∑–∏")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['phone_add'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–†–µ–±–µ–Ω–æ–∫ –±—É–¥–µ—Ç —É—Ö–æ–¥–∏—Ç—å —Å–∞–º –ø–æ –æ–∫–æ–Ω—á–∞–Ω–∏–∏ –¥–Ω—è?")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['leave'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ö—Ç–æ –∫—Ä–æ–º–µ —Ä–æ–¥–∏—Ç–µ–ª—è (–∑–∞–∫–æ–Ω–Ω–æ–≥–æ –ø—Ä–µ–¥—Å—Ç–∞–≤–∏—Ç–µ–ª—è) –º–æ–∂–µ—Ç –∑–∞–±–∏—Ä–∞—Ç—å "
                                                     "—Ä–µ–±–µ–Ω–∫–∞? –§–ò–û, –∫–µ–º –ø—Ä–∏—Ö–æ–¥–∏—Ç—Å—è, –∫–æ–Ω—Ç–∞–∫—Ç–Ω—ã–π —Ç–µ–ª–µ—Ñ–æ–Ω.")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['additional_contact'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ê–¥—Ä–µ—Å —Ñ–∞–∫—Ç–∏—á–µ—Å–∫–æ–≥–æ –ø—Ä–æ–∂–∏–≤–∞–Ω–∏—è —Ä–µ–±–µ–Ω–∫–∞")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['addr'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ù–æ–º–µ—Ä —Å—Ç—Ä–∞—Ö–æ–≤–æ–≥–æ –º–µ–¥–∏—Ü–∏–Ω—Å–∫–æ–≥–æ –ø–æ–ª–∏—Å–∞")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['oms'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–°–≤–µ–¥–µ–Ω–∏—è –æ —Å–æ—Å—Ç–æ—è–Ω–∏–∏ –∑–¥–æ—Ä–æ–≤—å—è —Ä–µ–±–µ–Ω–∫–∞:")
    cell.font = Font(name='TimesNewRoman', size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    ws.merge_cells(start_row=num_rows + 1, start_column=1, end_row=num_rows + 1, end_column=2)

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ï—Å—Ç—å –ª–∏ —É —Ä–µ–±–µ–Ω–∫–∞ –∑–∞–±–æ–ª–µ–≤–∞–Ω–∏—è (—Å–µ—Ä–¥—Ü–∞, –ø–∏—â–µ–≤–∞—Ä–∏—Ç–µ–ª—å–Ω–æ–π —Å–∏—Å—Ç–µ–º—ã, "
                                                     "–Ω–µ—Ä–≤–Ω–æ–π —Å–∏—Å—Ç–µ–º—ã, –ø—Å–∏—Ö–∏—á–µ—Å–∫–∏–µ —Ä–∞—Å—Å—Ç—Ä–æ–π—Å—Ç–≤–∞, –æ–ø–æ—Ä–Ω–æ-–¥–≤–∏–≥–∞—Ç–µ–ª—å–Ω–æ–π "
                                                     "—Å–∏—Å—Ç–µ–º—ã, –≤–Ω—É—Ç—Ä–µ–Ω–Ω–∏—Ö –æ—Ä–≥–∞–Ω–æ–≤). –ï—Å–ª–∏ –µ—Å—Ç—å, –ø–æ—è—Å–Ω–∏—Ç–µ –ø–æ–¥—Ä–æ–±–Ω–µ–µ.")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['disease'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ï—Å—Ç—å –ª–∏ —É —Ä–µ–±–µ–Ω–∫–∞ –∞–ª–ª–µ—Ä–≥–∏—á–µ—Å–∫–∞—è —Ä–µ–∞–∫—Ü–∏—è (–µ—Å–ª–∏ –µ—Å—Ç—å, —É–∫–∞–∂–∏—Ç–µ, "
                                                     "–Ω–∞ —á—Ç–æ –∏ –∫–∞–∫ –æ–Ω–∞ –ø—Ä–æ—è–≤–ª—è–µ—Ç—Å—è): ")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['allergy'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ë—ã–ª–∏ –ª–∏ —É —Ä–µ–±–µ–Ω–∫–∞ (–û–ø–µ—Ä–∞—Ü–∏–∏, –ü–µ—Ä–µ–ª–æ–º—ã, –°–æ—Ç—Ä—è—Å–µ–Ω–∏–µ –º–æ–∑–≥–∞, "
                                                     "–ü—Ä–∏—Å—Ç—É–ø—ã —ç–ø–∏–ª–µ–ø—Å–∏–∏). –ï—Å–ª–∏ –¥–∞, –ø–æ—è—Å–Ω–∏—Ç–µ –ø–æ–¥—Ä–æ–±–Ω–µ–µ.")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['other'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ï—Å—Ç—å –ª–∏ —É —Ä–µ–±–µ–Ω–∫–∞ –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏—è –ø–æ —Ñ–∏–∑–∏—á–µ—Å–∫–∏–º –Ω–∞–≥—Ä—É–∑–∫–∞–º? –ï—Å–ª–∏ –¥–∞, "
                                                     "–ø–æ—è—Å–Ω–∏—Ç–µ –ø–æ–¥—Ä–æ–±–Ω–µ–µ.")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['physic'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–µ —Å–≤–µ–¥–µ–Ω–∏—è")
    cell.font = Font(name='TimesNewRoman', size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    ws.merge_cells(start_row=num_rows + 1, start_column=1, end_row=num_rows + 1, end_column=2)

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ë—É–¥–µ—Ç –ª–∏ —Ä–µ–±–µ–Ω–æ–∫ –ø–æ—Å–µ—â–∞—Ç—å –±–∞—Å—Å–µ–π–Ω –≤–æ –≤—Ä–µ–º—è —Å–º–µ–Ω—ã?")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['swimm'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ù—É–∂–Ω—ã –ª–∏ —Ä–µ–±–µ–Ω–∫—É –≤ –±–∞—Å—Å–µ–π–Ω–µ –Ω–∞—Ä—É–∫–∞–≤–Ω–∏–∫–∏ –∏–ª–∏ –∂–∏–ª–µ—Ç?")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['jacket_swimm'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ß–µ–º —É–≤–ª–µ–∫–∞–µ—Ç—Å—è –í–∞—à —Ä–µ–±–µ–Ω–æ–∫? (—Ö–æ–±–±–∏, –ª—é–±–∏–º–æ–µ –∑–∞–Ω—è—Ç–∏–µ)")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['hobby'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ö–∞–∫–æ–µ —É—á–µ–±–Ω–æ–µ –∑–∞–≤–µ–¥–µ–Ω–∏–µ –ø–æ—Å–µ—â–∞–µ—Ç —Ä–µ–±–µ–Ω–æ–∫?")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['school'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ï—Å–ª–∏ –≤—ã —Ö–æ—Ç–∏—Ç–µ –µ—â–µ —á—Ç–æ-—Ç–æ —Ä–∞—Å—Å–∫–∞–∑–∞—Ç—å –Ω–∞–º –æ —Å–≤–æ–µ–º —Ä–µ–±–µ–Ω–∫–µ, "
                                                     "–Ω–∞–ø–∏—à–∏—Ç–µ —ç—Ç–æ –∑–¥–µ—Å—å")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['additional_info'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–ö–∞–∫ –í—ã —É–∑–Ω–∞–ª–∏ –ø—Ä–æ –∫–ª—É–± –ì–æ—Ä–æ–¥–æ–∫?")
    cell.font = Font(name='TimesNewRoman', size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['referer'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row

    cell = ws.cell(row=num_rows + 1, column=1, value="–í–∞—à email:")
    cell.font = Font(name='TimesNewRoman', size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    cell = ws.cell(row=num_rows + 1, column=2, value=df['email'][0])
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

    num_rows = ws.max_row + 1

    text = f"–Ø, {df['parent_main_name'][0]}, –∑–∞–∫–æ–Ω–Ω—ã–π –ø—Ä–µ–¥—Å—Ç–∞–≤–∏—Ç–µ–ª—å –Ω–µ—Å–æ–≤–µ—Ä—à–µ–Ω–Ω–æ–ª–µ—Ç–Ω–µ–≥–æ –≥—Ä–∞–∂–¥–∞–Ω–∏–Ω–∞"
    cell = ws.cell(row=num_rows + 1, column=1, value=text)
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=num_rows + 1, start_column=1, end_row=num_rows + 1, end_column=2)

    num_rows = ws.max_row + 1

    cell = ws.cell(row=num_rows, column=1,
                   value=f"___________________________________________ {df['child_birthday'][0].strftime('%d.%m.%Y')} –≥.—Ä.")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=num_rows, start_column=1, end_row=num_rows, end_column=2)

    num_rows = ws.max_row + 1

    cell = ws.cell(row=num_rows, column=1, value="–ü–æ–¥—Ç–≤–µ—Ä–∂–¥–∞—é –ø—Ä–µ–¥–æ—Å—Ç–∞–≤–ª–µ–Ω–Ω—ã–µ –º–Ω–æ–π —Å–≤–µ–¥–µ–Ω–∏—è")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    cell = ws.cell(row=num_rows, column=2, value="____________________")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    ws.row_dimensions[num_rows].height = 30

    num_rows = ws.max_row + 1

    cell = ws.cell(row=num_rows, column=2, value="(–ø–æ–¥–ø–∏—Å—å)")
    cell.font = Font(name='TimesNewRoman', size=8)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    num_rows = ws.max_row + 1

    cell = ws.cell(row=num_rows, column=1, value="–¥–∞—é —Å–≤–æ–µ —Å–æ–≥–ª–∞—Å–∏–µ –Ω–∞ –ø–æ–ª—É—á–µ–Ω–∏–µ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–æ–Ω–Ω—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π –æ—Ç –¥–µ—Ç—Å–∫–æ–≥–æ "
                                                 "–∫–ª—É–±–∞ –ì–æ—Ä–æ–¥–æ–∫")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    cell = ws.cell(row=num_rows, column=2, value="____________________")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    ws.row_dimensions[num_rows].height = 30

    num_rows = ws.max_row + 1

    cell = ws.cell(row=num_rows, column=2, value="(–ø–æ–¥–ø–∏—Å—å)")
    cell.font = Font(name='TimesNewRoman', size=8)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    num_rows = ws.max_row + 1

    cell = ws.cell(row=num_rows, column=1, value="–ø–æ–¥—Ç–≤–µ—Ä–∂–¥–∞—é —Å–≤–æ–µ —Å–æ–≥–ª–∞—Å–∏–µ –Ω–∞ –æ–±—Ä–∞–±–æ—Ç–∫—É —Å–≤–æ–∏—Ö –ø–µ—Ä—Å–æ–Ω–∞–ª—å–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö –∏"
                                                 "—Å–≤–æ–µ–≥–æ —Ä–µ–±–µ–Ω–∫–∞ ")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    cell = ws.cell(row=num_rows, column=2, value="____________________")
    cell.font = Font(name='TimesNewRoman', size=9)
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    ws.row_dimensions[num_rows].height = 30

    num_rows = ws.max_row + 1

    cell = ws.cell(row=num_rows, column=2, value="(–ø–æ–¥–ø–∏—Å—å)")
    cell.font = Font(name='TimesNewRoman', size=8)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    wb.save(output)
    output.seek(0)
    return output


if st.button("üìé –°–∫–∞—á–∞—Ç—å –∞–Ω–∫–µ—Ç—É –≤ .xlsx"):
    excel_file = create_ds(data)
    st.download_button(
        label="üíæ–ù–∞–∂–º–∏ –¥–ª—è —Å–∫–∞—á–∏–≤–∞–Ω–∏—è –∞–Ω–∫–µ—Ç—ã",
        data=excel_file,
        file_name=f"–ê–Ω–∫–µ—Ç–∞_{data['name'][0]}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
